"""
Enterprise Excel Report Generator for UniAttend Analytics.
Generates multi-sheet stylized workbooks with KPI summaries, student matrices,
conditional threshold formatting, and subject-wise defaulter rosters using openpyxl.
"""

import os
from datetime import datetime, date
from typing import Optional, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from database.models import Department, Course, Student, StudentCourseSummary, University, College
from database.db_manager import get_db_session

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output", "reports")


class ExcelAttendanceReporter:
    """Generates corporate-grade Excel reports for universities."""

    def __init__(self, output_dir: str = OUTPUT_DIR):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

        # Style definitions
        self.font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.font_title = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
        self.font_bold = Font(name="Calibri", size=10, bold=True)
        self.font_regular = Font(name="Calibri", size=10)
        
        # Fills
        self.fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Navy Blue
        self.fill_sub_header = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")  # Lighter Blue
        self.fill_green = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Safe
        self.fill_yellow = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Warning
        self.fill_red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Defaulter

        # Thin Border
        thin_side = Side(style="thin", color="CBD5E1")
        self.border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def generate_department_master_report(self, session: Session, department_id: int) -> str:
        """
        Generates a comprehensive multi-tab Excel report for a specific department:
          - Tab 1: Executive Summary
          - Tab 2: Master Student Roster
          - Tab 3: Defaulters Action List (<75%)
        """
        dept = session.get(Department, department_id)
        if not dept:
            raise ValueError(f"Department with ID {department_id} not found.")

        college = dept.college
        uni = college.university

        wb = Workbook()
        
        # ----------------------------------------------------
        # TAB 1: EXECUTIVE SUMMARY
        # ----------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary["A1"] = f"{uni.name} - {college.name}"
        ws_summary["A1"].font = self.font_title
        ws_summary["A2"] = f"Department of {dept.name} | Attendance Analytics Report"
        ws_summary["A2"].font = Font(name="Calibri", size=12, bold=True, color="475569")
        ws_summary["A3"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary["A3"].font = Font(name="Calibri", size=9, italic=True, color="64748B")

        # Fetch department metrics
        courses = session.query(Course).filter_by(department_id=dept.id).all()
        course_ids = [c.id for c in courses]
        
        summaries = session.query(StudentCourseSummary).filter(
            StudentCourseSummary.course_id.in_(course_ids)
        ).all()

        total_students = session.query(Student).filter_by(department_id=dept.id).count()
        total_enrollments = len(summaries)
        defaulter_count = len([s for s in summaries if s.is_defaulter])
        avg_pct = (sum(s.attendance_pct for s in summaries) / total_enrollments) if total_enrollments > 0 else 0.0

        # KPI Block
        ws_summary["A5"] = "KEY PERFORMANCE INDICATORS (KPIs)"
        ws_summary["A5"].font = self.font_bold
        
        kpi_headers = ["Metric", "Value", "Benchmark / Status"]
        for col_idx, h in enumerate(kpi_headers, 1):
            cell = ws_summary.cell(row=6, column=col_idx, value=h)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = Alignment(horizontal="center")

        kpi_data = [
            ("Total Enrolled Students", total_students, "Active"),
            ("Active Courses", len(courses), f"Semester {courses[0].semester if courses else 1}"),
            ("Average Institutional Attendance", f"{avg_pct:.2f}%", "Healthy" if avg_pct >= 78 else "Attention Required"),
            ("Total Defaulter Cases (< 75%)", defaulter_count, f"{(defaulter_count/total_enrollments*100):.1f}% of enrollments" if total_enrollments > 0 else "0%"),
            ("Attendance Eligibility Threshold", "75.0%", "Mandatory University Policy")
        ]

        for r_idx, (m, v, s) in enumerate(kpi_data, 7):
            ws_summary.cell(row=r_idx, column=1, value=m).font = self.font_regular
            c_val = ws_summary.cell(row=r_idx, column=2, value=v)
            c_val.font = self.font_bold
            c_val.alignment = Alignment(horizontal="center")
            ws_summary.cell(row=r_idx, column=3, value=s).font = self.font_regular
            for col in range(1, 4):
                ws_summary.cell(row=r_idx, column=col).border = self.border_thin

        # Course Level Summary Table
        ws_summary["A14"] = "COURSE-WISE ATTENDANCE PERFORMANCE"
        ws_summary["A14"].font = self.font_bold

        course_headers = ["Course Code", "Course Name", "Credits", "Total Classes", "Enrolled", "Defaulters", "Avg %"]
        for col_idx, h in enumerate(course_headers, 1):
            cell = ws_summary.cell(row=15, column=col_idx, value=h)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = Alignment(horizontal="center")

        for r_idx, c in enumerate(courses, 16):
            c_sums = [s for s in summaries if s.course_id == c.id]
            c_defaulters = len([s for s in c_sums if s.is_defaulter])
            c_avg = (sum(s.attendance_pct for s in c_sums) / len(c_sums)) if c_sums else 0.0
            c_total_cls = c_sums[0].total_classes if c_sums else 0

            ws_summary.cell(row=r_idx, column=1, value=c.course_code).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=r_idx, column=2, value=c.course_name)
            ws_summary.cell(row=r_idx, column=3, value=c.credits).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=r_idx, column=4, value=c_total_cls).alignment = Alignment(horizontal="center")
            ws_summary.cell(row=r_idx, column=5, value=len(c_sums)).alignment = Alignment(horizontal="center")
            
            c_def_cell = ws_summary.cell(row=r_idx, column=6, value=c_defaulters)
            c_def_cell.alignment = Alignment(horizontal="center")
            if c_defaulters > 0:
                c_def_cell.fill = self.fill_yellow

            c_pct_cell = ws_summary.cell(row=r_idx, column=7, value=f"{c_avg:.1f}%")
            c_pct_cell.alignment = Alignment(horizontal="center")
            c_pct_cell.font = self.font_bold
            if c_avg < 75.0:
                c_pct_cell.fill = self.fill_red

            for col in range(1, 8):
                ws_summary.cell(row=r_idx, column=col).border = self.border_thin

        # ----------------------------------------------------
        # TAB 2: MASTER STUDENT ROSTER
        # ----------------------------------------------------
        ws_roster = wb.create_sheet(title="Student Roster Matrix")
        ws_roster.views.sheetView[0].showGridLines = True

        roster_headers = [
            "Roll Number", "Student Name", "Course Code", "Course Name", 
            "Total Held", "Attended", "Late", "Absent", "Attendance %", "Status", "Risk Level"
        ]
        for col_idx, h in enumerate(roster_headers, 1):
            cell = ws_roster.cell(row=1, column=col_idx, value=h)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = Alignment(horizontal="center")

        students = session.query(Student).filter_by(department_id=dept.id).order_by(Student.student_id_str.asc()).all()
        student_dict = {s.id: s for s in students}
        course_dict = {c.id: c for c in courses}

        row_counter = 2
        for s_sum in summaries:
            st = student_dict.get(s_sum.student_id)
            cr = course_dict.get(s_sum.course_id)
            if not st or not cr:
                continue

            pct = s_sum.attendance_pct
            status_text = "ELIGIBLE" if pct >= 75.0 else "DEFAULTER"

            ws_roster.cell(row=row_counter, column=1, value=st.student_id_str)
            ws_roster.cell(row=row_counter, column=2, value=st.full_name)
            ws_roster.cell(row=row_counter, column=3, value=cr.course_code).alignment = Alignment(horizontal="center")
            ws_roster.cell(row=row_counter, column=4, value=cr.course_name)
            ws_roster.cell(row=row_counter, column=5, value=s_sum.total_classes).alignment = Alignment(horizontal="center")
            ws_roster.cell(row=row_counter, column=6, value=s_sum.attended_classes).alignment = Alignment(horizontal="center")
            ws_roster.cell(row=row_counter, column=7, value=s_sum.late_classes).alignment = Alignment(horizontal="center")
            ws_roster.cell(row=row_counter, column=8, value=s_sum.absent_classes).alignment = Alignment(horizontal="center")
            
            pct_cell = ws_roster.cell(row=row_counter, column=9, value=f"{pct:.1f}%")
            pct_cell.alignment = Alignment(horizontal="center")
            pct_cell.font = self.font_bold

            status_cell = ws_roster.cell(row=row_counter, column=10, value=status_text)
            status_cell.alignment = Alignment(horizontal="center")
            status_cell.font = self.font_bold

            risk_cell = ws_roster.cell(row=row_counter, column=11, value=s_sum.risk_category)
            risk_cell.alignment = Alignment(horizontal="center")

            # Apply conditional color styling
            if pct >= 80.0:
                pct_cell.fill = self.fill_green
                status_cell.fill = self.fill_green
            elif pct >= 75.0:
                pct_cell.fill = self.fill_yellow
                status_cell.fill = self.fill_yellow
            else:
                pct_cell.fill = self.fill_red
                status_cell.fill = self.fill_red

            for col in range(1, 12):
                ws_roster.cell(row=row_counter, column=col).border = self.border_thin

            row_counter += 1

        # ----------------------------------------------------
        # TAB 3: DEFAULTERS ACTION LIST (< 75%)
        # ----------------------------------------------------
        ws_defaulters = wb.create_sheet(title="Defaulters Action List")
        ws_defaulters.views.sheetView[0].showGridLines = True

        defaulter_headers = [
            "Roll Number", "Student Name", "Email", "Course Code", "Course Name", 
            "Held", "Attended", "Current %", "Shortage (Lectures Needed)", "Action Required"
        ]
        for col_idx, h in enumerate(defaulter_headers, 1):
            cell = ws_defaulters.cell(row=1, column=col_idx, value=h)
            cell.font = self.font_header
            cell.fill = PatternFill(start_color="B91C1C", end_color="B91C1C", fill_type="solid")  # Crimson Red
            cell.alignment = Alignment(horizontal="center")

        d_row = 2
        for s_sum in summaries:
            if not s_sum.is_defaulter:
                continue

            st = student_dict.get(s_sum.student_id)
            cr = course_dict.get(s_sum.course_id)
            if not st or not cr:
                continue

            ws_defaulters.cell(row=d_row, column=1, value=st.student_id_str)
            ws_defaulters.cell(row=d_row, column=2, value=st.full_name)
            ws_defaulters.cell(row=d_row, column=3, value=st.email)
            ws_defaulters.cell(row=d_row, column=4, value=cr.course_code).alignment = Alignment(horizontal="center")
            ws_defaulters.cell(row=d_row, column=5, value=cr.course_name)
            ws_defaulters.cell(row=d_row, column=6, value=s_sum.total_classes).alignment = Alignment(horizontal="center")
            ws_defaulters.cell(row=d_row, column=7, value=s_sum.attended_classes).alignment = Alignment(horizontal="center")
            
            d_pct = ws_defaulters.cell(row=d_row, column=8, value=f"{s_sum.attendance_pct:.1f}%")
            d_pct.alignment = Alignment(horizontal="center")
            d_pct.font = self.font_bold
            d_pct.fill = self.fill_red

            needed_cell = ws_defaulters.cell(row=d_row, column=9, value=f"+{s_sum.classes_needed_for_75} consecutive")
            needed_cell.alignment = Alignment(horizontal="center")
            needed_cell.font = self.font_bold

            action_cell = ws_defaulters.cell(row=d_row, column=10, value="Issue Warning Letter & Parent Intimation")
            action_cell.font = Font(name="Calibri", size=9, italic=True)

            for col in range(1, 11):
                ws_defaulters.cell(row=d_row, column=col).border = self.border_thin

            d_row += 1

        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    max_len = max(max_len, len(val_str))
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Save workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Attendance_Master_{dept.dept_code}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)

        print(f"[EXCEL REPORTER] Generated Master Excel report at: {filepath}")
        return filepath
