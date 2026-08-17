"""
Unit & Integration Tests for Automated Excel and PDF Report Generators.
"""

import pytest
import os
import sys
from openpyxl import load_workbook

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from database.db_manager import get_db_session, init_db
from database.models import Department, Student
from data.data_generator import seed_hierarchy_and_academics, generate_raw_attendance_stream
from pipeline.etl_pipeline import AttendanceETLPipeline
from reporting.excel_reporter import ExcelAttendanceReporter
from reporting.pdf_reporter import PDFReportGenerator
from reporting.automated_job import AutomatedReportingScheduler


@pytest.fixture(scope="module")
def prepared_db():
    init_db()
    with get_db_session() as session:
        seed_hierarchy_and_academics(session)
        generate_raw_attendance_stream(session, weeks=6)
    pipeline = AttendanceETLPipeline()
    pipeline.run_pipeline()
    yield


def test_excel_master_report_generation(prepared_db):
    reporter = ExcelAttendanceReporter()
    with get_db_session() as session:
        dept = session.query(Department).first()
        filepath = reporter.generate_department_master_report(session, dept.id)

        assert os.path.exists(filepath)
        assert os.path.getsize(filepath) > 1000

        # Validate sheets
        wb = load_workbook(filepath)
        assert "Executive Summary" in wb.sheetnames
        assert "Student Roster Matrix" in wb.sheetnames
        assert "Defaulters Action List" in wb.sheetnames


def test_pdf_executive_report_generation(prepared_db):
    pdf_gen = PDFReportGenerator()
    with get_db_session() as session:
        dept = session.query(Department).first()
        filepath = pdf_gen.generate_department_executive_pdf(session, dept.id)

        assert os.path.exists(filepath)
        assert os.path.getsize(filepath) > 2000


def test_pdf_student_warning_letter_generation(prepared_db):
    pdf_gen = PDFReportGenerator()
    with get_db_session() as session:
        student = session.query(Student).first()
        filepath = pdf_gen.generate_student_warning_letter_pdf(session, student.id)

        assert os.path.exists(filepath)
        assert os.path.getsize(filepath) > 1500


def test_automated_nightly_job(prepared_db):
    scheduler = AutomatedReportingScheduler()
    batch_res = scheduler.run_nightly_batch_job()

    assert batch_res["total_departments_audited"] > 0
    assert batch_res["excel_reports_generated"] > 0
    assert batch_res["pdf_executive_reports_generated"] > 0
